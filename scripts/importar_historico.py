"""
Import T_historico.xlsx + T_detalhes historico.xlsx
  -> consultas, evolucoes, odontograma (reusing existing prontuarios in estab=4)

Run: python scripts/importar_historico.py
"""
import sys, os, re, time
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import psycopg
from datetime import datetime, date

DATABASE_URL = os.environ.get(
    'DATABASE_URL',
    'postgresql://sisgersa:tOJ0rv1qWUQABYIWRMO0ew2c2AtfGZNU@dpg-d9hqikr7uimc73dt3e0g-a.oregon-postgres.render.com/sisgersa'
)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ESTAB_ID = 4


def main():
    t0 = time.time()
    conn = psycopg.connect(DATABASE_URL, sslmode='require', connect_timeout=30)
    cur = conn.cursor()

    print('[%.1f] Carregando mapas de lookup...' % (time.time()-t0))

    # Pacientes: codigo_paciente -> id
    cur.execute("SELECT id, codigo_paciente FROM pacientes WHERE codigo_paciente IS NOT NULL")
    pac_map = {str(row[1]): row[0] for row in cur.fetchall()}
    print('  pacientes: %d' % len(pac_map))

    # Profissionais: old -> new
    cur.execute("SELECT old_codigo, new_id FROM _old_prof_map")
    prof_map = {row[0]: row[1] for row in cur.fetchall()}
    print('  profissionais: %d' % len(prof_map))

    # Procedimentos: old -> new
    cur.execute("SELECT old_codigo, new_id FROM _old_proc_map")
    proc_map = {row[0]: row[1] for row in cur.fetchall()}
    print('  procedimentos: %d' % len(proc_map))

    # Prontuarios existentes em estab=4 (um por paciente)
    cur.execute("SELECT id, paciente_usuario_id FROM prontuarios WHERE estabelecimento_id = %s", (ESTAB_ID,))
    pront_by_paciente = {row[1]: row[0] for row in cur.fetchall()}
    print('  prontuarios em estab=%d: %d' % (ESTAB_ID, len(pront_by_paciente)))
    print('  estabelecimento_id: %d' % ESTAB_ID)

    # --- Carregar T_historico ---
    hist_file = 'T_hist\xf3rico.xlsx'
    print('\n[%.1f] Carregando %s...' % (time.time()-t0, hist_file))
    wb_h = openpyxl.load_workbook(os.path.join(BASE_DIR, hist_file), read_only=True, data_only=True)
    ws_h = wb_h[wb_h.sheetnames[0]]
    historicos = {}
    for row in ws_h.iter_rows(min_row=2, values_only=True):
        cod_hist, cod_cliente, alta, data_hist = (row + (None,) * 4)[:4]
        if cod_hist:
            hist_pac = pac_map.get(str(cod_cliente)) if cod_cliente else None
            if hist_pac:
                historicos[cod_hist] = {'paciente_id': hist_pac, 'alta': str(alta or '')}
    wb_h.close()
    print('  historicos validos: %d' % len(historicos))

    # --- Carregar T_detalhes historico ---
    det_file = 'T_detalhes hist\xf3rico.xlsx'
    print('[%.1f] Carregando %s...' % (time.time()-t0, det_file))
    wb_d = openpyxl.load_workbook(os.path.join(BASE_DIR, det_file), read_only=True, data_only=True)
    ws_d = wb_d[wb_d.sheetnames[0]]

    detalhes = []
    for row in ws_d.iter_rows(min_row=2, values_only=True):
        vals = (row + (None,) * 14)[:14]
        id_det, cod_hist, data_c, cod_p, cod_dent, obs, dente, prox_c, obs2, sel, busca, data_ret, id_odonto, proc_ant = vals
        if not cod_hist or cod_hist not in historicos:
            continue
        if not data_c:
            continue
        pac_id = historicos[cod_hist]['paciente_id']
        # Verificar se o paciente tem prontuario em estab=4
        if pac_id not in pront_by_paciente:
            continue
        detalhes.append({
            'cod_hist': cod_hist,
            'paciente_id': pac_id,
            'prontuario_id': pront_by_paciente[pac_id],
            'data_consulta': data_c,
            'cod_p': cod_p,
            'cod_dent': cod_dent,
            'obs': str(obs).strip() if obs else '',
            'dente': str(dente).strip() if dente else '',
            'prox_consulta': str(prox_c).strip() if prox_c else '',
        })
    wb_d.close()
    print('  detalhes validos: %d' % len(detalhes))

    if not detalhes:
        print('Nada a importar!')
        return

    # --- Inserir consultas + evolucoes + odontograma em batches ---
    print('\n[%.1f] Inserindo consultas e evolucoes...' % (time.time()-t0))

    now = datetime.now()
    consulta_count = 0
    evolucao_count = 0
    odonto_count = 0
    SUB_BATCH = 100

    def process_sub_batch(sub):
        nonlocal consulta_count, evolucao_count, odonto_count
        if not sub:
            return

        now2 = datetime.now()

        # --- Batch insert consultas with RETURNING ---
        cols_c = ('paciente_usuario_id', 'profissional_usuario_id', 'estabelecimento_id',
                   'prontuario_id', 'procedimento_id', 'data_hora', 'status', 'observacoes',
                   'criado_em', 'atualizado_em')
        vals_c = []
        params_c = []
        for item in sub:
            vals_c.append('(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)')
            params_c.extend([
                item['paciente_id'], item['prof_id'], ESTAB_ID,
                item['prontuario_id'], item['proc_id'], item['data_hora'],
                'concluida', item['obs'], now2, now2
            ])
        sql_c = 'INSERT INTO consultas (%s) VALUES %s RETURNING id' % (
            ', '.join(cols_c), ', '.join(vals_c)
        )
        cur.execute(sql_c, params_c)
        consulta_ids = [r[0] for r in cur.fetchall()]
        consulta_count += len(consulta_ids)

        # --- Batch insert evolucoes ---
        cols_e = ('prontuario_id', 'consulta_id', 'profissional_usuario_id',
                   'data', 'procedimento_realizado', 'observacoes', 'criado_em')
        vals_e = []
        params_e = []
        for idx, item in enumerate(sub):
            vals_e.append('(%s, %s, %s, %s, %s, %s, %s)')
            params_e.extend([
                item['prontuario_id'], consulta_ids[idx], item['prof_id'],
                item['data_hora'].date(), None, item['obs'], now2
            ])
        sql_e = 'INSERT INTO evolucoes (%s) VALUES %s' % (
            ', '.join(cols_e), ', '.join(vals_e)
        )
        cur.execute(sql_e, params_e)
        evolucao_count += len(sub)

        # --- Batch insert odontograma ---
        odonto_items = [item for item in sub if item.get('dente')]
        if odonto_items:
            cols_o = ('prontuario_id', 'dente', 'face', 'condicao',
                       'data_registro', 'profissional_usuario_id', 'criado_em')
            vals_o = []
            params_o = []
            for item in odonto_items:
                vals_o.append('(%s, %s, %s, %s, %s, %s, %s)')
                params_o.extend([
                    item['prontuario_id'], item['dente'], None, 'tratado',
                    item['data_hora'].date(), item['prof_id'], now2
                ])
            sql_o = 'INSERT INTO odontograma (%s) VALUES %s' % (
                ', '.join(cols_o), ', '.join(vals_o)
            )
            cur.execute(sql_o, params_o)
            odonto_count += len(odonto_items)

        conn.commit()

    sub = []
    for i, d in enumerate(detalhes):
        data_c = d['data_consulta']
        obs = d['obs']
        dente_str = d['dente']
        prox = d['prox_consulta']

        if isinstance(data_c, datetime):
            data_hora = data_c
        elif isinstance(data_c, date):
            data_hora = datetime.combine(data_c, datetime.min.time())
        else:
            continue

        obs_final = obs
        if prox:
            obs_final = (obs_final + '\nProx: ' + prox) if obs_final else 'Prox: ' + prox

        dente_val = None
        if dente_str:
            try:
                dn = int(re.sub(r'\D', '', dente_str))
                if 1 <= dn <= 52:
                    dente_val = dn
            except ValueError:
                pass

        sub.append({
            'paciente_id': d['paciente_id'],
            'prof_id': prof_map.get(d['cod_dent'], 1),
            'prontuario_id': d['prontuario_id'],
            'proc_id': proc_map.get(d['cod_p']) if d['cod_p'] else None,
            'data_hora': data_hora,
            'obs': obs_final or None,
            'dente': dente_val,
        })

        if len(sub) >= SUB_BATCH:
            process_sub_batch(sub)
            sub = []
            if (i + 1) % 1000 == 0:
                print('  %d/%d processados...' % (i+1, len(detalhes)))

    if sub:
        process_sub_batch(sub)
    print('  %d/%d processados' % (len(detalhes), len(detalhes)))

    print('\n[%.1f] Concluido!' % (time.time()-t0))
    print('  Consultas: %d' % consulta_count)
    print('  Evolucoes: %d' % evolucao_count)
    print('  Odontograma: %d' % odonto_count)

    cur.close()
    conn.close()


if __name__ == '__main__':
    main()
