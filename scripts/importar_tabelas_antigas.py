"""
Import script: T_procedimento.xlsx -> procedimentos
              T_profissionais.xlsx -> profissionais (com mapeamento _old_prof_map)

Run: python scripts/importar_tabelas_antigas.py
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import psycopg
import bcrypt
from datetime import datetime
from utils.db_url import get_database_url

DATABASE_URL = get_database_url()
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def parse_duracao(texto):
    if not texto:
        return None
    texto = str(texto).strip().lower()
    total = 0
    h = re.search(r'(\d+)\s*hora', texto)
    if h:
        total += int(h.group(1)) * 60
    m = re.search(r'(\d+)\s*minuto', texto)
    if m:
        total += int(m.group(1))
    return total if total > 0 else None


def importar_procedimentos(cur):
    print('\n=== IMPORTANDO PROCEDIMENTOS ===')
    wb = openpyxl.load_workbook(
        os.path.join(BASE_DIR, 'T_procedimento.xlsx'),
        read_only=True, data_only=True
    )
    ws = wb[wb.sheetnames[0]]

    mapa_categoria = {
        1: 'cirurgia', 2: 'radiologia', 3: 'cirurgia',
        4: 'endodontia', 5: 'protese', 6: 'protese',
        7: 'periodontia', 8: 'cirurgia'
    }

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (row + (None,) * 15)[:15]
        if not vals[1]:
            continue

        duracao = parse_duracao(vals[9])
        area = vals[3]
        categoria = mapa_categoria.get(area) if area else None
        copasa = vals[4]
        codigo_tuss = str(int(copasa)) if copasa and str(copasa).strip() else None
        nome_copasa = str(vals[6]).strip() if vals[6] else None

        nome_proc = str(vals[1]).strip()
        cur.execute("SELECT id FROM procedimentos WHERE nome = %s", (nome_proc,))
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO procedimentos (nome, descricao, duracao_minutos, ativo, categoria, codigo_tuss, criado_em, atualizado_em)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (nome_proc, nome_copasa, duracao, True, categoria, codigo_tuss, datetime.now(), datetime.now()))
            inserted += 1

    wb.close()
    print(f'  {inserted} procedimentos inseridos')


CARGO_MAP = {
    'dentista': 'dentista', 'cd': 'dentista',
    'medico': 'medico', 'dr': 'medico',
    'enfermeiro': 'enfermeiro',
    'admin': 'admin', 'administrador': 'admin', 'administrador(a)': 'admin',
    'recepcionista': 'recepcionista', 'secretaria': 'recepcionista', 'secretario': 'recepcionista',
    'auxiliar': 'auxiliar', 'auxiliar de consultorio odontologico': 'auxiliar',
    'auxiliar de consultório odontológico': 'auxiliar',
    'atendente': 'auxiliar',
    'periodontista': 'dentista', 'periodontia': 'dentista',
    'endodontista': 'dentista', 'endodontia': 'dentista',
    'ortodontia': 'dentista', 'ortodontista': 'dentista',
    'clinico': 'dentista', 'clinica': 'dentista', 'clínico': 'dentista',
    'odontopediatria': 'dentista', 'pediatria': 'dentista',
    'protese': 'dentista', 'prótese': 'dentista',
    'cirurgia': 'dentista', 'cirurgia e traumatologia buco-maxilo-facial': 'dentista',
}


def normalizar_cargo(texto):
    if not texto:
        return None
    t = texto.strip().lower()
    # Direct mapping
    if t in CARGO_MAP:
        return CARGO_MAP[t]
    # Check if any keyword is in the text
    for keyword, mapped in sorted(CARGO_MAP.items(), key=lambda x: -len(x[0])):
        if keyword in t:
            return mapped
    return None


def importar_profissionais(cur):
    print('\n=== IMPORTANDO PROFISSIONAIS ===')
    wb = openpyxl.load_workbook(
        os.path.join(BASE_DIR, 'T_profissionais.xlsx'),
        read_only=True, data_only=True
    )
    ws = wb[wb.sheetnames[0]]

    cur.execute("""
        CREATE TABLE IF NOT EXISTS _old_prof_map (
            old_codigo INTEGER PRIMARY KEY,
            new_id INTEGER NOT NULL REFERENCES profissionais(id)
        )
    """)
    cur.execute("DELETE FROM _old_prof_map")

    inserted = 0
    updated = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (row + (None,) * 24)[:24]
        codigo, nome = vals[0], vals[2]
        telefone_r, telefone_c, telefone_cel, cro, cpf = vals[8], vals[9], vals[10], vals[11], vals[12]
        especialidade, categoria_d = vals[13], vals[16]
        obs, data_nasc, ativo = vals[18], vals[19], vals[21]

        if not nome or not str(nome).strip():
            skipped += 1
            continue

        nome = str(nome).strip()
        telefone = None
        for t in [telefone_cel, telefone_c, telefone_r]:
            if t and str(t).strip():
                tel = re.sub(r'\D', '', str(t))
                if len(tel) >= 10:
                    telefone = tel
                    break

        if not telefone:
            skipped += 1
            continue

        # Extrair email do campo observacao
        email_base = None
        if obs and '@' in str(obs):
            m = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', str(obs))
            if m:
                email_base = m.group(0)
        if not email_base:
            email_base = f'profissional_{codigo}@sisgersa.local'

        cpf_clean = re.sub(r'\D', '', str(cpf)) if cpf else None
        if cpf_clean and len(cpf_clean) != 11:
            cpf_clean = None

        cat = str(categoria_d).strip() if categoria_d else ''
        is_dentista = cat == '1'
        is_recepcionista = cat == '6'

        cargo = normalizar_cargo(especialidade)
        if is_recepcionista and not cargo:
            cargo = 'recepcionista'
        if is_dentista and not cargo:
            cargo = 'dentista'

        ativo_bool = True
        if ativo is not None:
            if isinstance(ativo, bool):
                ativo_bool = ativo
            elif isinstance(ativo, (int, float)):
                ativo_bool = ativo == 1
            elif str(ativo).strip() in ('0', 'False', 'false'):
                ativo_bool = False

        senha_hash = bcrypt.hashpw(b'senha123', bcrypt.gensalt()).decode('utf-8')

        # Check if exists by email (no unique constraint on email, but we use it as key)
        cur.execute("SELECT id FROM profissionais WHERE email = %s", (email_base,))
        existing = cur.fetchone()
        if existing:
            cur.execute("""
                UPDATE profissionais SET nome=%s, telefone=%s, cpf=%s, cargo=%s, ativo=%s,
                    is_dentista=%s, is_recepcionista=%s, atualizado_em=%s
                WHERE id=%s
            """, (nome, telefone, cpf_clean, cargo, ativo_bool,
                  is_dentista, is_recepcionista, datetime.now(), existing[0]))
            prof_id = existing[0]
            updated += 1
        else:
            cur.execute("""
                INSERT INTO profissionais (nome, email, senha_hash, telefone, cpf, cargo, ativo,
                    is_dentista, is_recepcionista, criado_em, atualizado_em)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (nome, email_base, senha_hash, telefone, cpf_clean, cargo, ativo_bool,
                  is_dentista, is_recepcionista, datetime.now(), datetime.now()))
            cur.execute("SELECT id FROM profissionais WHERE email = %s", (email_base,))
            r = cur.fetchone()
            if r:
                prof_id = r[0]
            else:
                continue
            inserted += 1

        cur.execute("""
            INSERT INTO _old_prof_map (old_codigo, new_id) VALUES (%s, %s)
            ON CONFLICT (old_codigo) DO UPDATE SET new_id = EXCLUDED.new_id
        """, (codigo, prof_id))

    wb.close()
    print(f'  {inserted} inseridos, {updated} atualizados, {skipped} ignorados')
    cur.execute("SELECT count(*) FROM _old_prof_map")
    print(f'  {cur.fetchone()[0]} mapeamentos salvos em _old_prof_map')


def main():
    print('Conectando ao banco...')
    conn = psycopg.connect(DATABASE_URL, sslmode='require')
    cur = conn.cursor()
    try:
        importar_procedimentos(cur)
        importar_profissionais(cur)
        conn.commit()
        print('\nImportacao concluida com sucesso!')
    except Exception as e:
        conn.rollback()
        print(f'\nERRO: {e}')
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    main()
