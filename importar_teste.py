import openpyxl
import psycopg
import hashlib
from datetime import datetime
import sys
import traceback

PG_URL = 'postgresql://sisgersa:tOJ0rv1qWUQABYIWRMO0ew2c2AtfGZNU@dpg-d9hqikr7uimc73dt3e0g-a.oregon-postgres.render.com/sisgersa'
ESTAB_ID = 4
DEFAULT_SENHA_HASH = hashlib.sha256('sisgersa123'.encode()).hexdigest()

print("1. Lendo Excel...")
sys.stdout.flush()
wb = openpyxl.load_workbook('T_cliente.xlsx')
ws = wb['T_cliente']
print(f"   OK: {ws.max_row} linhas")
sys.stdout.flush()

headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
print(f"   Headers: {headers[:10]}...")
sys.stdout.flush()

print("2. Conectando ao PostgreSQL...")
sys.stdout.flush()
conn = psycopg.connect(PG_URL)
conn.autocommit = False
cur = conn.cursor()
print("   OK")
sys.stdout.flush()

def safe_str(val):
    if val is None: return None
    s = str(val).strip()
    return s if s and s != 'None' else None

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    s = str(val).strip()
    if not s or s == 'None': return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

print("3. Iniciando importacao...")
sys.stdout.flush()
inserted = 0
skipped = 0
errors = 0

for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=200, values_only=True), start=2):
    try:
        data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}

        nome = safe_str(data.get('NOME'))
        if not nome: skipped += 1; continue

        sobrenome = safe_str(data.get('SOBRENOME'))
        nome_completo = f"{nome} {sobrenome}" if sobrenome else nome
        codigo_c = safe_str(data.get('Codigo_c'))
        if codigo_c: codigo_c = str(codigo_c).strip()

        email = safe_str(data.get('E_mail_c'))
        if not email and codigo_c:
            email = f"paciente_{codigo_c}@sisgersa.local"
        elif not email:
            email = f"paciente_imp_{row_idx}@sisgersa.local"

        cpf_raw = safe_str(data.get('CPF_c'))
        cpf = None
        if cpf_raw:
            cpf = cpf_raw.replace('.', '').replace('-', '').replace('/', '')
            if len(cpf) < 10: cpf = None

        telefone = safe_str(data.get('TelefoneCel_c')) or safe_str(data.get('TelefoneR_c'))
        data_nasc = parse_date(data.get('Data de nascimento'))

        endereco_parts = []
        for f in ['Logadouro_c', 'BAIRRO', 'CIDADE', 'ESTADO']:
            v = safe_str(data.get(f))
            if v: endereco_parts.append(v)
        endereco = ', '.join(endereco_parts) if endereco_parts else None

        indicacao = None
        for k in headers:
            if 'indica' in k.lower():
                indicacao = safe_str(data.get(k))
                break

        estado_civil = safe_str(data.get('ESTADO CIVIL'))
        profissao = None
        for k in headers:
            if 'profiss' in k.lower():
                profissao = safe_str(data.get(k))
                break

        nome_pai = safe_str(data.get('NOME DO PAI'))
        nome_mae = safe_str(data.get('NOME DA MAE'))

        num_doc = None
        for k in headers:
            if 'numero' in k.lower() and 'documenta' in k.lower():
                num_doc = safe_str(data.get(k))
                break

        if codigo_c:
            cur.execute("SELECT id FROM usuarios WHERE codigo_paciente = %s", (codigo_c,))
            if cur.fetchone():
                skipped += 1; continue

        cur.execute("""
            INSERT INTO usuarios (nome, email, senha_hash, tipo, telefone, cpf, data_nascimento,
                                  endereco, codigo_paciente, numero_documentacao, indicacao,
                                  estado_civil, profissao, nome_pai, nome_mae, ativo)
            VALUES (%s, %s, %s, 'paciente', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (nome_completo, email, DEFAULT_SENHA_HASH, telefone, cpf, data_nasc,
              endereco, codigo_c, num_doc, indicacao, estado_civil, profissao, nome_pai, nome_mae))
        user_id = cur.fetchone()[0]

        cur.execute("""
            INSERT INTO paciente_estabelecimento (usuario_id, estabelecimento_id, data_cadastro)
            VALUES (%s, %s, COALESCE(%s, CURRENT_DATE))
        """, (user_id, ESTAB_ID, data_nasc))

        inserted += 1
        if inserted % 50 == 0:
            conn.commit()
            print(f"   ... {inserted} inseridos (linha {row_idx})")
            sys.stdout.flush()

    except Exception as e:
        errors += 1
        if errors <= 5:
            print(f"ERRO linha {row_idx}: {e}")
            traceback.print_exc()
            sys.stdout.flush()

conn.commit()
print(f"\n=== TESTE (200 linhas) ===")
print(f"Inseridos: {inserted}")
print(f"Pulados: {skipped}")
print(f"Erros: {errors}")
cur.close(); conn.close()
