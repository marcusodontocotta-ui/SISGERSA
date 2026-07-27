import openpyxl
import psycopg
from datetime import datetime
import sys

PG_URL = 'postgresql://sisgersa:tOJ0rv1qWUQABYIWRMO0ew2c2AtfGZNU@dpg-d9hqikr7uimc73dt3e0g-a.oregon-postgres.render.com/sisgersa'

print("1. Lendo Excel...")
sys.stdout.flush()
wb = openpyxl.load_workbook('T_cliente.xlsx')
ws = wb['T_cliente']
headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
print(f"   Headers relevantes:")
for i, h in enumerate(headers):
    if 'digo' in h.lower() or 'docum' in h.lower() or 'indica' in h.lower():
        print(f"     [{i}] '{h}'")
sys.stdout.flush()

def safe(val):
    if val is None: return None
    s = str(val).strip()
    return s if s and s != 'None' else None

def parse_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    s = str(val).strip()
    if not s or s == 'None': return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

print("2. Conectando...")
sys.stdout.flush()
conn = psycopg.connect(PG_URL)
conn.autocommit = False
cur = conn.cursor()

cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND codigo_paciente IS NULL")
total_sem_codigo = cur.fetchone()[0]
print(f"   {total_sem_codigo} pacientes sem codigo_paciente")
sys.stdout.flush()

print("3. Atualizando codigo_paciente e numero_documentacao...")
sys.stdout.flush()

updated = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    try:
        d = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}

        codigo_c = None
        for k in headers:
            if 'digo_c' in k.lower():
                codigo_c = safe(d.get(k))
                break
        if codigo_c:
            codigo_c = str(codigo_c).strip()

        num_doc = None
        for k in headers:
            if 'numero' in k.lower() and 'documenta' in k.lower():
                num_doc = safe(d.get(k))
                break

        if not codigo_c and not num_doc:
            continue

        nome = safe(d.get('NOME'))
        sobrenome = safe(d.get('SOBRENOME'))
        nome_completo = f"{nome} {sobrenome}" if sobrenome else nome
        if not nome_completo:
            continue

        if codigo_c:
            cur.execute("SELECT id FROM usuarios WHERE codigo_paciente = %s", (codigo_c,))
            existing = cur.fetchone()
            if existing:
                if num_doc:
                    cur.execute("UPDATE usuarios SET numero_documentacao = %s WHERE codigo_paciente = %s AND numero_documentacao IS NULL",
                                (num_doc, codigo_c))
                continue

        cur.execute("""
            UPDATE usuarios SET codigo_paciente = %s, numero_documentacao = %s
            WHERE tipo = 'paciente' AND nome = %s AND codigo_paciente IS NULL
            RETURNING id
        """, (codigo_c, num_doc, nome_completo))
        result = cur.fetchone()

        if result:
            updated += 1
        else:
            print(f"   Nao encontrado: {nome_completo} (codigo={codigo_c})")

        if updated % 200 == 0 and updated > 0:
            conn.commit()
            print(f"   ... {updated} atualizados (linha {row_idx})")
            sys.stdout.flush()

    except Exception as e:
        print(f"ERRO linha {row_idx}: {e}")
        sys.stdout.flush()

conn.commit()
print(f"\n=== RESULTADO ===")
print(f"Atualizados com codigo: {updated}")

cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND codigo_paciente IS NOT NULL")
print(f"Total com codigo_paciente: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND numero_documentacao IS NOT NULL")
print(f"Total com numero_documentacao: {cur.fetchone()[0]}")

cur.close(); conn.close()
