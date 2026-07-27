import openpyxl
import psycopg
import hashlib
from datetime import datetime
import sys

PG_URL = 'postgresql://sisgersa:tOJ0rv1qWUQABYIWRMO0ew2c2AtfGZNU@dpg-d9hqikr7uimc73dt3e0g-a.oregon-postgres.render.com/sisgersa'
ESTAB_ID = 4
DEFAULT_HASH = hashlib.sha256('sisgersa123'.encode()).hexdigest()
BATCH = 300

print("1. Lendo Excel...")
sys.stdout.flush()
wb = openpyxl.load_workbook('T_cliente.xlsx')
ws = wb['T_cliente']
headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

COL = {}
for i, h in enumerate(headers):
    hl = h.lower()
    if 'digo_c' in hl: COL['codigo'] = i
    elif h.strip() == 'NOME': COL['nome'] = i
    elif h.strip() == 'SOBRENOME': COL['sobrenome'] = i
    elif 'nascimento' in hl: COL['data_nasc'] = i
    elif 'logadouro' in hl: COL['logradouro'] = i
    elif h.strip() == 'BAIRRO': COL['bairro'] = i
    elif h.strip() == 'CIDADE': COL['cidade'] = i
    elif h.strip() == 'ESTADO': COL['estado'] = i
    elif h.strip() == 'CEP_c': COL['cep'] = i
    elif 'cel' in hl and 'telefone' in hl: COL['tel_cel'] = i
    elif h.strip() == 'TelefoneR_c' and 'tel_res' not in COL: COL['tel_res'] = i
    elif 'indica' in hl and 'indicacao' not in COL: COL['indicacao'] = i
    elif h.strip() == 'CPF_c': COL['cpf'] = i
    elif 'e_mail' in hl: COL['email'] = i
    elif 'nome do pai' in hl: COL['nome_pai'] = i
    elif 'nome da mae' in hl: COL['nome_mae'] = i
    elif 'estado civil' in hl: COL['estado_civil'] = i
    elif 'profiss' in hl and 'profissao' not in COL: COL['profissao'] = i
    elif 'numero' in hl and 'documenta' in hl: COL['num_doc'] = i
    elif 'convenio' in hl and 'convenio' not in COL: COL['convenio'] = i

print(f"   Mapeamento: {COL}")
print(f"   Total linhas: {ws.max_row}")
sys.stdout.flush()

def s(val):
    if val is None: return None
    v = str(val).strip()
    return v if v and v != 'None' else None

def pd(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    s2 = str(val).strip()
    if not s2 or s2 == 'None': return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try: return datetime.strptime(s2, fmt).date()
        except: continue
    return None

def g(idx):
    return COL.get(idx)

print("2. Conectando...")
sys.stdout.flush()
conn = psycopg.connect(PG_URL)
conn.autocommit = False
cur = conn.cursor()

print("3. Importando...")
sys.stdout.flush()
inserted = 0
skipped = 0
errors = 0
all_rows = []

for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        nome = s(row[g('nome')]) if g('nome') is not None else None
        if not nome: skipped += 1; continue

        sobrenome = s(row[g('sobrenome')]) if g('sobrenome') is not None else None
        nome_completo = f"{nome} {sobrenome}" if sobrenome else nome

        codigo = s(row[g('codigo')]) if g('codigo') is not None else None
        if codigo: codigo = str(codigo).strip()

        email = s(row[g('email')]) if g('email') is not None else None
        if not email and codigo: email = f"paciente_{codigo}@sisgersa.local"
        elif not email: skipped += 1; continue

        cpf = None
        cpf_raw = s(row[g('cpf')]) if g('cpf') is not None else None
        if cpf_raw:
            cpf = cpf_raw.replace('.', '').replace('-', '').replace('/', '')
            if len(cpf) < 10: cpf = None

        tel = None
        if g('tel_cel') is not None: tel = s(row[g('tel_cel')])
        if not tel and g('tel_res') is not None: tel = s(row[g('tel_res')])

        dn = pd(row[g('data_nasc')]) if g('data_nasc') is not None else None

        ep = []
        for f in ['logradouro', 'bairro', 'cidade', 'estado']:
            if g(f) is not None:
                v = s(row[g(f)])
                if v: ep.append(v)
        endereco = ', '.join(ep) if ep else None

        indicacao = s(row[g('indicacao')]) if g('indicacao') is not None else None
        ec = s(row[g('estado_civil')]) if g('estado_civil') is not None else None
        prof = s(row[g('profissao')]) if g('profissao') is not None else None
        pai = s(row[g('nome_pai')]) if g('nome_pai') is not None else None
        mae = s(row[g('nome_mae')]) if g('nome_mae') is not None else None
        ndoc = s(row[g('num_doc')]) if g('num_doc') is not None else None

        all_rows.append((nome_completo, email, DEFAULT_HASH, tel, cpf, dn,
                          endereco, codigo, ndoc, indicacao, ec, prof, pai, mae))
    except Exception as e:
        errors += 1
        if errors <= 3: print(f"ERRO parse: {e}")

print(f"   {len(all_rows)} pacientes para inserir ({skipped} pulados)")
sys.stdout.flush()

print("   Verificando duplicados no banco...")
sys.stdout.flush()
cur.execute("SELECT codigo_paciente FROM usuarios WHERE codigo_paciente IS NOT NULL")
existing_codes = {r[0] for r in cur.fetchall()}
print(f"   {len(existing_codes)} codigos ja existentes")
sys.stdout.flush()

to_insert = [r for r in all_rows if not r[7] or r[7] not in existing_codes]
print(f"   {len(to_insert)} realmente para inserir")
sys.stdout.flush()

print("   Buscando ultimo numero de prontuario...")
sys.stdout.flush()
cur.execute("SELECT COUNT(*) FROM prontuarios WHERE estabelecimento_id = %s", (ESTAB_ID,))
pront_counter = cur.fetchone()[0]
print(f"   Prontuarios existentes: {pront_counter}")
sys.stdout.flush()

for i in range(0, len(to_insert), BATCH):
    batch = to_insert[i:i+BATCH]
    try:
        user_ids = []
        for r in batch:
            cur.execute("""
                INSERT INTO usuarios (nome, email, senha_hash, tipo, telefone, cpf, data_nascimento,
                                      endereco, codigo_paciente, numero_documentacao, indicacao,
                                      estado_civil, profissao, nome_pai, nome_mae, ativo)
                VALUES (%s,%s,%s,'paciente',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
                RETURNING id
            """, r)
            user_ids.append(cur.fetchone()[0])

        pe_args = [(uid, ESTAB_ID, batch[j][5] or datetime.now().date()) for j, uid in enumerate(user_ids)]
        cur.executemany(
            "INSERT INTO paciente_estabelecimento (usuario_id, estabelecimento_id, data_cadastro) VALUES (%s,%s,%s)",
            pe_args)

        pr_args = []
        for uid in user_ids:
            pront_counter += 1
            pr_args.append((uid, ESTAB_ID, f"PRONT-{pront_counter:05d}"))
        cur.executemany(
            "INSERT INTO prontuarios (paciente_usuario_id, estabelecimento_id, numero_prontuario) VALUES (%s,%s,%s)",
            pr_args)

        conn.commit()
        inserted += len(batch)
        print(f"   ... {inserted}/{len(to_insert)}")
        sys.stdout.flush()

    except Exception as e:
        conn.rollback()
        errors += len(batch)
        print(f"ERRO batch {i}: {e}")
        sys.stdout.flush()

print(f"\n=== RESULTADO ===")
print(f"Inseridos: {inserted}")
print(f"Pulados: {skipped}")
print(f"Erros: {errors}")

cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND codigo_paciente IS NOT NULL")
print(f"Com codigo_paciente: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND numero_documentacao IS NOT NULL")
print(f"Com numero_documentacao: {cur.fetchone()[0]}")

cur.close(); conn.close()
