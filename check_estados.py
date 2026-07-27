import openpyxl
import sys

wb = openpyxl.load_workbook('T_cliente.xlsx')
ws = wb['T_cliente']
headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

estado_idx = None
for i, h in enumerate(headers):
    if h.strip() == 'ESTADO':
        estado_idx = i
        break

estados = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if estado_idx is not None:
        v = row[estado_idx]
        if v:
            estados.add(str(v).strip())

print("Valores unicos de estado:")
for e in sorted(estados):
    print(f"  '{e}' ({len(e)} chars)")
