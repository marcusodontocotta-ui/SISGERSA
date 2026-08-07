import openpyxl
import psycopg
import sys
from utils.db_url import get_database_url

PG_URL = get_database_url()
BATCH = 300

print("1. Lendo Excel...")
sys.stdout.flush()
wb = openpyxl.load_workbook('T_cliente.xlsx')
ws = wb['T_cliente']
headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

COL = {}
for i, h in enumerate(headers):
    hl = h.lower()
    if 'digo_c' in hl: COL['codigo'] = i
    elif h.strip() == 'NOME': COL['nome'] = i
    elif h.strip() == 'SOBRENOME': COL['sobrenome'] = i
    elif 'logadouro' in hl: COL['logradouro'] = i
    elif h.strip() == 'BAIRRO': COL['bairro'] = i
    elif h.strip() == 'CIDADE': COL['cidade'] = i
    elif h.strip() == 'ESTADO': COL['estado'] = i
    elif h.strip() == 'CEP_c': COL['cep'] = i
    elif h.strip() == 'Numero_c': COL['numero'] = i

def s(val):
    if val is None: return None
    v = str(val).strip()
    return v if v and v != 'None' else None

def g(idx):
    return COL.get(idx)

print("2. Conectando...")
sys.stdout.flush()
conn = psycopg.connect(PG_URL)
conn.autocommit = False
cur = conn.cursor()

print("3. Buscando pacientes existentes por codigo_paciente...")
sys.stdout.flush()
cur.execute("SELECT id, codigo_paciente FROM usuarios WHERE tipo='paciente' AND codigo_paciente IS NOT NULL")
code_to_id = {r[1]: r[0] for r in cur.fetchall()}
print(f"   {len(code_to_id)} pacientes com codigo encontrado")
sys.stdout.flush()

print("4. Atualizando enderecos...")
sys.stdout.flush()
updated = 0
skipped = 0
errors = 0
updates = []

for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        codigo = s(row[g('codigo')]) if g('codigo') is not None else None
        if not codigo: skipped += 1; continue
        codigo = str(codigo).strip()
        if codigo not in code_to_id: skipped += 1; continue

        logradouro = s(row[g('logradouro')]) if g('logradouro') is not None else None
        bairro = s(row[g('bairro')]) if g('bairro') is not None else None
        cidade = s(row[g('cidade')]) if g('cidade') is not None else None
        estado = s(row[g('estado')]) if g('estado') is not None else None
        cep = s(row[g('cep')]) if g('cep') is not None else None
        if cep: cep = str(cep).replace('.', '').replace('-', '').strip()
        numero = s(row[g('numero')]) if g('numero') is not None else None

        if logradouro or bairro or cidade or estado or cep:
            updates.append((logradouro, numero, bairro, cidade, estado, cep, code_to_id[codigo]))
    except Exception as e:
        errors += 1
        if errors <= 3: print(f"ERRO parse: {e}")

print(f"   {len(updates)} para atualizar ({skipped} pulados)")
sys.stdout.flush()

for i in range(0, len(updates), BATCH):
    batch = updates[i:i+BATCH]
    try:
        cur.executemany(
            """UPDATE usuarios SET logradouro = %s, numero = %s, bairro = %s,
               cidade = %s, estado = %s, cep = %s WHERE id = %s""",
            batch
        )
        conn.commit()
        updated += len(batch)
        print(f"   ... {updated}/{len(updates)}")
        sys.stdout.flush()
    except Exception as e:
        conn.rollback()
        errors += len(batch)
        print(f"ERRO batch {i}: {e}")
        sys.stdout.flush()

print(f"\n=== RESULTADO ===")
print(f"Atualizados: {updated}")
print(f"Pulados: {skipped}")
print(f"Erros: {errors}")

cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND logradouro IS NOT NULL")
print(f"Com logradouro: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND cep IS NOT NULL")
print(f"Com cep: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM usuarios WHERE tipo='paciente' AND cidade IS NOT NULL")
print(f"Com cidade: {cur.fetchone()[0]}")

cur.close(); conn.close()
